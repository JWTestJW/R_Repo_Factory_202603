#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Read and validate Excel parameters from NewRepoInfo.xlsx
"""

import os
import re
import sys
import json
from openpyxl import load_workbook


# 許可文字
ALLOWED_VALUE_PATTERN = r"^[A-Za-z0-9_-]+$"
ALLOWED_VALUE_DESC = "letters, digits, underscore (_), and hyphen (-)"


def strip_spaces(value):
    """文字列の前後にある半角/全角スペースを除去する。"""
    if value is None:
        return ""
    
    # 文字列に変換する
    value_str = str(value).strip()
    
    # 全角スペース (U+3000) を除去する
    value_str = value_str.strip('\u3000')
    
    return value_str


def write_has_errors_output(has_errors):
    """has_errors を GitHub Actions 出力へ書き込む。"""
    output_path = os.getenv('GITHUB_OUTPUT')
    if not output_path:
        return

    with open(output_path, 'a', encoding='utf-8') as output_file:
        output_file.write(f"has_errors={'true' if has_errors else 'false'}\n")


def write_json_output(data_list):
    """json_output を GitHub Actions 出力へ書き込む。"""
    output_path = os.getenv('GITHUB_OUTPUT')
    if not output_path:
        return

    json_single_line = json.dumps(data_list, ensure_ascii=False)
    with open(output_path, 'a', encoding='utf-8') as output_file:
        output_file.write(f"json_output<<EOF\n{json_single_line}\nEOF\n")


def validate_row(row_num, row_data_by_col, col_to_header, errors):
    """
    1行分のデータを検証する。

    引数:
    - row_num: 行番号
    - row_data_by_col: B列からG列までの値を持つ辞書（キーは列記号）
    - col_to_header: 列記号とヘッダー名の対応辞書
    - errors: エラーメッセージ格納用リスト

    戻り値:
    - 検証成功時はTrue、失敗時はFalse
    """
    columns = ['B', 'C', 'D', 'E', 'F', 'G']
    is_valid = True
    
    # D列の値を取得する
    d_value = row_data_by_col.get('D', '')
    
    # 各列を検証する
    for col in columns:
        value = row_data_by_col.get(col, '')
        header_name = col_to_header.get(col, col)

        # 値がある場合のみ、許可文字ルールを検証する
        if value and not re.fullmatch(ALLOWED_VALUE_PATTERN, value):
            errors.append(
                f"Row {row_num}, Column {header_name}: contains invalid characters. Allowed: {ALLOWED_VALUE_DESC}."
            )
            is_valid = False
        
        # F列（StudyNo）のみ特別ルールを適用する
        if col == 'F':
            if d_value == 'Study':
                # D列（リポジトリタイプ）がStudyの場合、F列（StudyNo）は必須
                if not value or value == '-':
                    d_header = col_to_header.get('D', 'D')
                    errors.append(
                        f"Row {row_num}, Column {header_name}: must not be empty or '-' when {d_header} is Study."
                    )
                    is_valid = False
        else:
            # それ以外の列は必須
            if not value:
                errors.append(f"Row {row_num}, Column {header_name}: must not be empty.")
                is_valid = False
    
    return is_valid


def main():
    """メイン処理。"""
    # 環境変数から設定値を取得する
    excel_file = os.getenv('EXCEL_FILE', 'NewRepoInfo.xlsx')
    sheet_name = os.getenv('SHEET_NAME', 'Prameters')
    
    print(f"Reading file: {excel_file}")
    print(f"Sheet name: {sheet_name}")
    
    # 対象ファイルの存在を確認する
    if not os.path.exists(excel_file):
        print(f"Error: File '{excel_file}' does not exist.")
        write_has_errors_output(True)
        sys.exit(1)
    
    try:
        # ワークブックを読み込む
        wb = load_workbook(excel_file, data_only=True)
        
        # 指定シートの存在を確認する
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' does not exist.")
            print(f"Available sheets: {', '.join(wb.sheetnames)}")
            write_has_errors_output(True)
            sys.exit(1)
        
        ws = wb[sheet_name]

        # B列からG列のヘッダーを1行目から取得する
        columns = ['B', 'C', 'D', 'E', 'F', 'G']
        col_to_header = {}
        for col in columns:
            header_value = strip_spaces(ws[f'{col}1'].value)
            col_to_header[col] = header_value if header_value else col
        
        # データ格納リストとエラー格納リスト
        data_list = []
        errors = []
        
        print("File content reading start:")

        # 4行目から読み取りを開始する
        row_num = 4
        while True:
            # B列を読み取る
            b_cell = ws[f'B{row_num}']
            b_value = strip_spaces(b_cell.value)
            
            # B列が空または空白のみの場合は読み取り終了
            if not b_value:
                if row_num == 4:
                    print("The file content is empty. Please check the file.")
                    errors.append("File content is empty.")
                else:
                    print("File content reading is complete.")
                break
            
            # B列からG列までを読み取る
            row_data_by_col = {
                'B': strip_spaces(ws[f'B{row_num}'].value),
                'C': strip_spaces(ws[f'C{row_num}'].value),
                'D': strip_spaces(ws[f'D{row_num}'].value),
                'E': strip_spaces(ws[f'E{row_num}'].value),
                'F': strip_spaces(ws[f'F{row_num}'].value),
                'G': strip_spaces(ws[f'G{row_num}'].value),
            }

            display_row_num = row_num - 3
            raw_row_text = ', '.join(row_data_by_col[col] for col in columns)
            print(f"Row {display_row_num}: {raw_row_text}")

            row_data = {col_to_header[col]: row_data_by_col[col] for col in columns}
            
            # 行データを検証し、結果に関わらず確認用に保持する
            validate_row(row_num, row_data_by_col, col_to_header, errors)
            data_list.append(row_data)
            
            row_num += 1
        
        print(f"Total rows read: {len(data_list)}")
        
        # エラー有無で処理を分岐する
        if errors:
            print("\nValidation errors found:")
            for error in errors:
                print(f"  - {error}")
            
            # GitHub Actions出力を設定する
            write_has_errors_output(True)
            
            sys.exit(1)
        else:
            print("\nAll data passed validation.")
            
            # 読み取り結果をJSONへ変換する
            json_output = json.dumps(data_list, ensure_ascii=False, indent=2)
            print("\nJSON output:")
            print(json_output)
            
            # GitHub Actions出力を設定する
            write_has_errors_output(False)
            write_json_output(data_list)
            
            print("\nThe has_errors and json_output values were successfully exported to GitHub Actions.")
    
    except Exception as e:
        print(f"An error occurred during processing: {str(e)}")
        import traceback
        traceback.print_exc()
        write_has_errors_output(True)
        sys.exit(1)


if __name__ == '__main__':
    main()
